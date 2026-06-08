import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
hf = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
hfill = PatternFill(start_color="1F4E79", fill_type="solid")
sfill = PatternFill(start_color="2E75B6", fill_type="solid")
pc = ["4472C4","70AD47","FFC000","ED7D31","C00000"]
afill = PatternFill(start_color="D6E4F0", fill_type="solid")
thin = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
ctr = Alignment(horizontal="center",vertical="center",wrap_text=True)
la = Alignment(horizontal="left",vertical="center",wrap_text=True)

def hdr(ws,r,c,val,fill=hfill,span=None):
    cell = ws.cell(row=r,column=c,value=val)
    cell.font=hf; cell.fill=fill; cell.alignment=ctr; cell.border=thin
    if span: ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+span-1)

def dat(ws,r,c,val,alt=False,fmt=None,bold=False):
    cell = ws.cell(row=r,column=c,value=val)
    cell.border=thin; cell.alignment=la if isinstance(val,str) else ctr
    if alt: cell.fill=afill
    if fmt: cell.number_format=fmt
    if bold: cell.font=Font(bold=True,name="Calibri",size=10)
    return cell

def sw(ws,widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w

wh = ["WBS ID","Task Name","Duration (Days)","Start","End","Predecessor","Assigned To","% Complete","Status","Deliverable","RAG","Notes"]

# ===== SHEET 1: DASHBOARD =====
ws=wb.active; ws.title="Project Dashboard"; ws.sheet_properties.tabColor="1F4E79"
ws.merge_cells('A1:O1'); ws['A1']="HONG KONG CAD RADAR PROJECT — PROJECT MONITORING DASHBOARD"
ws['A1'].font=Font(bold=True,size=16,color="FFFFFF",name="Calibri"); ws['A1'].fill=PatternFill(start_color="1F4E79",fill_type="solid"); ws['A1'].alignment=ctr
ws.merge_cells('A2:O2'); ws['A2']="Civil Aviation Department | 5-Phase Radar System Implementation | PMP Compliant"
ws['A2'].font=Font(size=11,color="1F4E79",name="Calibri",italic=True); ws['A2'].alignment=ctr
r=4
hdr(ws,r,1,"PROJECT INFORMATION",span=8); r+=1
for i,(k,v) in enumerate([("Project Name","HK CAD Radar System Modernization"),("Client","CAD,HK"),("PM","TBD"),("Code","CAD-RADAR-2025"),("Value","HK$850M"),("Start","2025-07-01"),("End","2029-06-30"),("Status","In Progress")]):
    dat(ws,r,i+1,k,bold=True); dat(ws,r+1,i+1,v)
r+=3; hdr(ws,r,1,"PHASE OVERVIEW",span=10); r+=1
for i,h in enumerate(["Phase","Name","Start","End","Days","Budget (HK$)","Actual (HK$)","% Done","Status","RAG"],1):
    hdr(ws,r,i,h,sfill)
r+=1
for pi,ph in enumerate([["1","Initiation","2025-07-01","2025-12-31","184","42.5M","38.2M","95%","Done","Green"],["2","Design","2026-01-01","2026-12-31","365","170M","95M","60%","Active","Green"],["3","Installation","2027-01-01","2027-12-31","365","340M","0","0%","Pending","Grey"],["4","Testing","2028-01-01","2028-12-31","365","127.5M","0","0%","Pending","Grey"],["5","Handover","2029-01-01","2029-06-30","182","170M","0","0%","Pending","Grey"]]):
    for ci,val in enumerate(ph,1):
        cell=dat(ws,r+pi,ci,val,alt=(pi%2==1))
        if ci==1: cell.fill=PatternFill(start_color=pc[pi],fill_type="solid"); cell.font=Font(bold=True,color="FFFFFF",size=10)
r+=6; hdr(ws,r,1,"KEY PERFORMANCE INDICATORS",span=10); r+=1
for i,h in enumerate(["KPI","Target","Actual","Variance","Unit","Status","Trend","PMP Area","Frequency","Owner"],1):
    hdr(ws,r,i,h,sfill)
r+=1
for ki,row in enumerate([["SPI","1.00","0.98","-0.02","Ratio","On Track","Up","Schedule","Weekly","PM"],["CPI","1.00","1.05","+0.05","Ratio","Favorable","Up","Cost","Weekly","PM"],["Scope Changes","<10","3","-7","Count","Favorable","Down","Scope","Monthly","PM"],["Defect Density","<0.5","0.3","-0.2","Def/KLOC","Favorable","Down","Quality","Monthly","QA"],["Risk Score","<50","35","-15","Score","OK","Flat","Risk","Monthly","RM"],["Stakeholder Sat",">80%","85%","+5%","%","Favorable","Up","Stakeholder","Quarterly","PM"],["Safety Incidents","0","0","0","Count","Perfect","Flat","Safety","Weekly","SO"],["Milestone On-Time",">90%","92%","+2%","%","Favorable","Up","Schedule","Monthly","PM"],["BAC","HK$850M","HK$850M","0","HK$","On Budget","Flat","Cost","Monthly","PM"],["EAC","HK$850M","HK$809M","-41M","HK$","Favorable","Up","Cost","Monthly","PM"]]):
    for ci,val in enumerate(row,1): dat(ws,r+ki,ci,val,alt=(ki%2==1))
r+=12; hdr(ws,r,1,"EARNED VALUE MANAGEMENT (EVM)",span=10); r+=1
for i,h in enumerate(["Metric","PV","EV","AC","SV","CV","SPI","CPI"],1): hdr(ws,r,i,h,sfill)
r+=1
for ei,ev in enumerate([["Cumulative","212.5M","208.25M","198.33M","-4.25M","+9.92M","0.980","1.050"],["Current Period","17M","16.66M","15.88M","-0.34M","+0.78M","0.980","1.050"],["EAC","—","—","—","—","809M","—","—"],["VAC","—","—","—","+41M","—","—","—"],["TCPI","—","—","—","—","—","—","1.02"]]):
    for ci,val in enumerate(ev,1): dat(ws,r+ei,ci,val,alt=(ei%2==1))
sw(ws,[22,20,16,16,16,18,18,12,12,8,14,14,14,14,16])
ws.freeze_panes="A5"
print("Sheet 1: Dashboard DONE")

# ===== SHEET 2: PHASE 1 =====
ws2=wb.create_sheet("Phase 1 - Initiation"); ws2.sheet_properties.tabColor=pc[0]
ws2.merge_cells('A1:L1'); ws2['A1']="PHASE 1: INITIATION & PLANNING"
ws2['A1'].font=Font(bold=True,size=14,color="FFFFFF",name="Calibri"); ws2['A1'].fill=PatternFill(start_color=pc[0],fill_type="solid"); ws2['A1'].alignment=ctr
r=3; hdr(ws2,r,1,"WORK BREAKDOWN STRUCTURE",span=12); r+=1
for i,h in enumerate(wh,1): hdr(ws2,r,i,h,PatternFill(start_color=pc[0],fill_type="solid"))
r+=1
wbs1=[["1.1","Project Charter Dev","14","2025-07-01","2025-07-14","—","Sponsor","100%","Done","Charter","Green","Approved"],["1.2","Stakeholder ID","21","2025-07-01","2025-07-21","—","PM","100%","Done","Register","Green","47 identified"],["1.3","Requirements Gathering","42","2025-07-15","2025-08-25","1.1","BA Team","100%","Done","Req Doc","Green","215 reqs"],["1.4","Feasibility Study","30","2025-07-22","2025-08-20","1.2","Eng Lead","100%","Done","Report","Green","Confirmed"],["1.5","Site Survey","28","2025-08-01","2025-08-28","1.2","Site Eng","100%","Done","Survey Report","Green","3 sites"],["1.6","Environmental Impact","35","2025-08-15","2025-09-18","1.4","Env Cons","100%","Done","EIA Report","Green","EPD approved"],["1.7","Risk Assessment","28","2025-08-26","2025-09-22","1.3,1.4","Risk Mgr","100%","Done","Risk Register","Green","68 risks"],["1.8","Budget Estimation","21","2025-09-01","2025-09-21","1.4","Cost Eng","100%","Done","Cost Est","Green","HK$850M"],["1.9","Project Plan Dev","30","2025-09-15","2025-10-14","1.3,1.7","PM","100%","Done","PMP","Green","PMBOK 7th"],["1.10","Regulatory Approval","45","2025-09-23","2025-11-06","1.6,1.8","PM","100%","Done","CAD Approval","Green","Obtained"],["1.11","Phase Gate Review","7","2025-11-15","2025-11-21","1.9,1.10","Steering","100%","Done","Gate Approval","Green","Authorized"],["1.12","Lessons Learned","10","2025-11-22","2025-12-01","1.11","PM","90%","In Progress","LL Report","Green","Draft done"]]
for wi,rd in enumerate(wbs1):
    for ci,val in enumerate(rd,1): dat(ws2,r+wi,ci,val,alt=(wi%2==1))
sw(ws2,[10,32,14,12,12,12,16,12,14,22,8,20]); ws2.freeze_panes="A4"
print("Sheet 2: Phase 1 DONE")

# ===== SHEET 3: PHASE 2 =====
ws3=wb.create_sheet("Phase 2 - Design"); ws3.sheet_properties.tabColor=pc[1]
ws3.merge_cells('A1:L1'); ws3['A1']="PHASE 2: DESIGN & PROCUREMENT"
ws3['A1'].font=Font(bold=True,size=14,color="FFFFFF",name="Calibri"); ws3['A1'].fill=PatternFill(start_color=pc[1],fill_type="solid"); ws3['A1'].alignment=ctr
r=3; hdr(ws3,r,1,"WORK BREAKDOWN STRUCTURE",span=12); r+=1
for i,h in enumerate(wh,1): hdr(ws3,r,i,h,PatternFill(start_color=pc[1],fill_type="solid"))
r+=1
wbs2=[["2.1","System Architecture","60","2026-01-01","2026-03-01","1.11","Architect","85%","In Progress","Arch Doc","Green","On schedule"],["2.2","Radar Tech Specs","45","2026-01-15","2026-02-28","2.1","Tech Lead","75%","In Progress","Tech Spec","Green","ASR-12"],["2.3","SSR Design","45","2026-02-01","2026-03-17","2.1","Comms Eng","65%","In Progress","SSR Doc","Amber","Mode S delay"],["2.4","MLAT Design","40","2026-02-15","2026-03-26","2.1","Nav Eng","50%","In Progress","MLAT Doc","Green","4 stations"],["2.5","Data Processing Design","55","2026-03-01","2026-04-24","2.1","SW Arch","40%","In Progress","SW Doc","Green","Asterix"],["2.6","Site Prep Design","50","2026-03-15","2026-05-03","2.2,2.3","Civil Eng","30%","In Progress","Plans","Green","3 sites"],["2.7","RFP Preparation","35","2026-04-01","2026-05-05","2.2","Procurement","25%","In Progress","RFP Docs","Green","3 packages"],["2.8","Tender Evaluation","60","2026-05-06","2026-07-04","2.7","Board","0%","Pending","Eval Report","Grey",""],["2.9","Contract Award","30","2026-07-05","2026-08-03","2.8","Legal","0%","Pending","Contracts","Grey",""],["2.10","FAT Plan","28","2026-06-01","2026-06-28","2.2,2.3","QA","0%","Pending","FAT Plan","Grey",""],["2.11","SAT Plan","28","2026-06-15","2026-07-12","2.4,2.5","QA","0%","Pending","SAT Plan","Grey",""],["2.12","PDR Review","14","2026-04-25","2026-05-08","2.1,2.2","CAD/PM","0%","Pending","PDR Minutes","Grey",""],["2.13","CDR Review","14","2026-08-04","2026-08-17","2.9,2.12","CAD/PM","0%","Pending","CDR Minutes","Grey",""],["2.14","Phase Gate Review","7","2026-11-15","2026-11-21","All","Steering","0%","Pending","Approval","Grey",""]]
for wi,rd in enumerate(wbs2):
    for ci,val in enumerate(rd,1): dat(ws3,r+wi,ci,val,alt=(wi%2==1))
r+=16; hdr(ws3,r,1,"PROCUREMENT TRACKER",span=12); r+=1
for i,h in enumerate(["Package","Description","RFP Date","Tender Close","Value (HK$)","Vendor","Status","RAG"],1):
    hdr(ws3,r,i,h,PatternFill(start_color=pc[1],fill_type="solid"))
r+=1
for pi,row in enumerate([["PKG-001","Primary Radar","2026-04-15","2026-05-30","280M","TBD (Thales)","RFP Issued","Green"],["PKG-002","SSR","2026-04-15","2026-05-30","120M","TBD (Raytheon)","RFP Issued","Green"],["PKG-003","MLAT","2026-05-01","2026-06-15","45M","TBD (ERA)","Draft RFP","Amber"],["PKG-004","Data Processing","2026-05-15","2026-06-30","95M","TBD (Indra)","Planning","Grey"],["PKG-005","Comms Infra","2026-06-01","2026-07-15","35M","TBD (Nokia)","Planning","Grey"],["PKG-006","Civil Works","2026-07-01","2026-08-15","65M","TBD (Local)","Planning","Grey"],["PKG-007","Installation Svc","2026-08-01","2026-09-15","85M","TBD (SI)","Planning","Grey"]]):
    for ci,val in enumerate(row,1): dat(ws3,r+pi,ci,val,alt=(pi%2==1))
sw(ws3,[10,28,12,12,14,20,14,8]); ws3.freeze_panes="A4"
print("Sheet 3: Phase 2 DONE")

# ===== SHEET 4: PHASE 3 =====
ws4=wb.create_sheet("Phase 3 - Installation"); ws4.sheet_properties.tabColor=pc[2]
ws4.merge_cells('A1:L1'); ws4['A1']="PHASE 3: SITE INSTALLATION & SYSTEM INTEGRATION"
ws4['A1'].font=Font(bold=True,size=14,color="000000",name="Calibri"); ws4['A1'].fill=PatternFill(start_color=pc[2],fill_type="solid"); ws4['A1'].alignment=ctr
r=3; hdr(ws4,r,1,"WORK BREAKDOWN STRUCTURE",span=12); r+=1
for i,h in enumerate(wh,1): hdr(ws4,r,i,h,PatternFill(start_color=pc[2],fill_type="solid"))
r+=1
wbs3=[["3.1","Site Prep - Site A","90","2027-01-15","2027-04-14","2.6","Contractor","0%","Pending","Site Ready","Grey","Chek Lap Kok"],["3.2","Site Prep - Site B","90","2027-01-15","2027-04-14","2.6","Contractor","0%","Pending","Site Ready","Grey","Tai Mo Shan"],["3.3","Site Prep - Site C","75","2027-02-01","2027-04-16","2.6","Contractor","0%","Pending","Site Ready","Grey","MLS"],["3.4","Install Primary Radar","120","2027-04-15","2027-08-12","3.1,PKG1","Vendor","0%","Pending","Install Cert","Grey",""],["3.5","Install SSR","90","2027-05-01","2027-07-29","3.2,PKG2","Vendor","0%","Pending","Install Cert","Grey",""],["3.6","Install MLAT (x4)","150","2027-04-16","2027-09-12","3.3,PKG3","Vendor","0%","Pending","Install Cert","Grey",""],["3.7","Install Data Proc","60","2027-06-01","2027-07-30","3.1,PKG4","Vendor","0%","Pending","Install Cert","Grey",""],["3.8","Network Install","75","2027-05-15","2027-07-28","3.1,PKG5","Vendor","0%","Pending","Install Cert","Grey",""],["3.9","System Integration","90","2027-08-13","2027-11-10","3.4-3.8","SI Lead","0%","Pending","Integ Report","Grey",""],["3.10","CAD Interface","60","2027-09-01","2027-10-30","3.7,3.8","SW Team","0%","Pending","Interface Cert","Grey","Indra system"],["3.11","Power Supply","45","2027-07-01","2027-08-14","3.1","Electrical","0%","Pending","Power Cert","Grey","UPS+Generator"],["3.12","Earthing & Lightning","30","2027-07-15","2027-08-13","3.1","Electrical","0%","Pending","Safety Cert","Grey",""],["3.13","Phase Gate Review","7","2027-11-20","2027-11-26","3.9-3.12","CAD/PM","0%","Pending","Approval","Grey",""]]
for wi,rd in enumerate(wbs3):
    for ci,val in enumerate(rd,1): dat(ws4,r+wi,ci,val,alt=(wi%2==1))
sw(ws4,[10,32,14,12,12,14,14,12,14,20,8,20]); ws4.freeze_panes="A4"
print("Sheet 4: Phase 3 DONE")

# ===== SHEET 5: PHASE 4 =====
ws5=wb.create_sheet("Phase 4 - Testing"); ws5.sheet_properties.tabColor=pc[3]
ws5.merge_cells('A1:L1'); ws5['A1']="PHASE 4: TESTING & COMMISSIONING"
ws5['A1'].font=Font(bold=True,size=14,color="000000",name="Calibri"); ws5['A1'].fill=PatternFill(start_color=pc[3],fill_type="solid"); ws5['A1'].alignment=ctr
r=3; hdr(ws5,r,1,"WORK BREAKDOWN STRUCTURE",span=12); r+=1
for i,h in enumerate(wh,1): hdr(ws5,r,i,h,PatternFill(start_color=pc[3],fill_type="solid"))
r+=1
wbs4=[["4.1","FAT (Factory Acceptance)","45","2028-01-15","2028-02-27","3.9","QA/Vendor","0%","Pending","FAT Report","Grey","All subsystems"],["4.2","SAT - Site A","30","2028-03-01","2028-03-30","4.1","QA Team","0%","Pending","SAT Report A","Grey",""],["4.3","SAT - Site B","30","2028-03-15","2028-04-13","4.1","QA Team","0%","Pending","SAT Report B","Grey",""],["4.4","SAT - Site C","25","2028-04-01","2028-04-25","4.1","QA Team","0%","Pending","SAT Report C","Grey",""],["4.5","End-to-End Test","30","2028-05-01","2028-05-30","4.2-4.4","QA/PM","0%","Pending","E2E Report","Grey","Full chain"],["4.6","OAT (Operational)","45","2028-06-01","2028-07-15","4.5","CAD Ops","0%","Pending","OAT Report","Grey","ATCO involved"],["4.7","Safety Case","30","2028-07-01","2028-07-30","4.6","Safety Eng","0%","Pending","Safety Case","Grey",""],["4.8","Regulatory Commissioning","30","2028-08-01","2028-08-30","4.7","CAD","0%","Pending","Commission Cert","Grey",""],["4.9","Parallel Running","90","2028-09-01","2028-11-29","4.8","Ops Team","0%","Pending","Parallel Report","Grey","3 months"],["4.10","Performance Validation","30","2028-11-01","2028-11-30","4.9","QA/PM","0%","Pending","Perf Report","Grey","ICAO stds"],["4.11","Phase Gate Review","7","2028-12-05","2028-12-11","4.10","Steering","0%","Pending","Approval","Grey",""]]
for wi,rd in enumerate(wbs4):
    for ci,val in enumerate(rd,1): dat(ws5,r+wi,ci,val,alt=(wi%2==1))
r+=13; hdr(ws5,r,1,"DEFECT TRACKER",span=12); r+=1
for i,h in enumerate(["Defect ID","Description","Severity","Reported","Status","Owner","Resolution","Phase Found"],1):
    hdr(ws5,r,i,h,PatternFill(start_color=pc[3],fill_type="solid"))
r+=1
for di,row in enumerate([["DEF-001","Radar display latency >2s","Major","2026-03-15","Resolved","SW Team","Algorithm fix","Phase 2"],["DEF-002","SSR Mode S interrogation fail","Critical","2026-05-20","In Progress","Comms Eng","HW replacement","Phase 2"],["DEF-003","MLAT sync drift","Minor","2026-06-10","Open","Nav Eng","GPS fix","Phase 2"]]):
    for ci,val in enumerate(row,1): dat(ws5,r+di,ci,val,alt=(di%2==1))
sw(ws5,[12,35,12,12,14,14,20,14]); ws5.freeze_panes="A4"
print("Sheet 5: Phase 4 DONE")

# ===== SHEET 6: PHASE 5 =====
ws6=wb.create_sheet("Phase 5 - Handover"); ws6.sheet_properties.tabColor=pc[4]
ws6.merge_cells('A1:L1'); ws6['A1']="PHASE 5: HANDOVER & PROJECT CLOSEOUT"
ws6['A1'].font=Font(bold=True,size=14,color="FFFFFF",name="Calibri"); ws6['A1'].fill=PatternFill(start_color=pc[4],fill_type="solid"); ws6['A1'].alignment=ctr
r=3; hdr(ws6,r,1,"WORK BREAKDOWN STRUCTURE",span=12); r+=1
for i,h in enumerate(wh,1): hdr(ws6,r,i,h,PatternFill(start_color=pc[4],fill_type="solid"))
r+=1
wbs5=[["5.1","System Handover","14","2029-01-01","2029-01-14","4.11","PM/Ops","0%","Pending","Handover Cert","Grey",""],["5.2","Documentation Handover","21","2029-01-08","2029-01-28","5.1","Tech Writer","0%","Pending","Doc Register","Grey","O&M manuals"],["5.3","Training - ATCO","30","2029-01-15","2029-02-13","5.1","Trainer","0%","Pending","Training Cert","Grey","40 ATCOs"],["5.4","Training - Maint Staff","21","2029-01-29","2029-02-18","5.1","Vendor","0%","Pending","Training Cert","Grey","25 engineers"],["5.5","Training - Eng Staff","14","2029-02-19","2029-03-04","5.1","Vendor","0%","Pending","Training Cert","Grey","15 engineers"],["5.6","Warranty Period Start","182","2029-01-15","2029-07-15","5.1","PM","0%","Pending","Warranty Cert","Grey","24 months"],["5.7","Final Documentation","30","2029-02-01","2029-03-02","5.2","PM","0%","Pending","Final Package","Grey","As-built"],["5.8","Financial Closeout","21","2029-03-15","2029-04-04","5.7","Finance","0%","Pending","Final Account","Grey","All payments"],["5.9","Lessons Learned Workshop","7","2029-04-05","2029-04-11","5.8","PM","0%","Pending","LL Report","Grey","All phases"],["5.10","Final Audit","14","2029-04-12","2029-04-25","5.9","Auditor","0%","Pending","Audit Report","Grey","Independent"],["5.11","Project Closure","7","2029-06-24","2029-06-30","5.10","PM","0%","Pending","Closure Report","Grey",""]]
for wi,rd in enumerate(wbs5):
    for ci,val in enumerate(rd,1): dat(ws6,r+wi,ci,val,alt=(wi%2==1))
sw(ws6,[10,32,14,12,12,12,14,12,14,22,8,20]); ws6.freeze_panes="A4"
print("Sheet 6: Phase 5 DONE")

# ===== SHEET 7: KPI & PMP SUMMARY =====
ws7=wb.create_sheet("KPI & PMP Summary")
ws7.merge_cells('A1:K1'); ws7['A1']="CONSOLIDATED KPI & PMP MEASUREMENTS"
ws7['A1'].font=Font(bold=True,size=14,color="FFFFFF",name="Calibri"); ws7['A1'].fill=hfill; ws7['A1'].alignment=ctr
r=3; hdr(ws7,r,1,"PMP KNOWLEDGE AREA COMPLIANCE",span=11); r+=1
for i,h in enumerate(["Knowledge Area","Process Group","Key Deliverable","Status","Compliance %","Auditor Notes"],1): hdr(ws7,r,i,h,sfill)
r+=1
for pi,row in enumerate([["Integration","All Phases","Project Mgmt Plan","On Track","95%","PMBOK 7"],["Scope","Planning","WBS / Reqs Doc","On Track","98%","215 reqs"],["Schedule","Plan/Monitor","Gantt / Milestones","On Track","92%","12 milestones"],["Cost","Plan/Monitor","Budget / EVM","On Track","96%","CPI=1.05"],["Quality","Executing","QA Plan / Defects","On Track","94%","ISO 9001"],["Resource","Executing","Resource Histogram","On Track","90%","52 FTEs"],["Communications","All","Comms Plan / Reports","On Track","95%","Weekly"],["Risk","Plan/Monitor","Risk Register","On Track","92%","68 risks"],["Procurement","Plan/Exec","Contracts / POs","On Track","88%","7 packages"],["Stakeholder","All","Stakeholder Register","On Track","97%","47 stakeholders"]]):
    for ci,val in enumerate(row,1): dat(ws7,r+pi,ci,val,alt=(pi%2==1))
r+=12; hdr(ws7,r,1,"EARNED VALUE MANAGEMENT DETAIL",span=11); r+=1
for i,h in enumerate(["Month","PV","EV","AC","SV","CV","SPI","CPI"],1): hdr(ws7,r,i,h,sfill)
r+=1
months_data=[["Jul-25",3.5,3.3,3.2,-0.2,0.1,0.943,1.031],["Aug-25",7.1,6.8,6.5,-0.3,0.3,0.958,1.046],["Sep-25",12.8,12.2,11.5,-0.6,0.7,0.953,1.061],["Oct-25",18.5,17.8,16.8,-0.7,1.0,0.962,1.060],["Nov-25",25.0,24.0,22.8,-1.0,1.2,0.960,1.053],["Dec-25",32.5,31.0,29.5,-1.5,1.5,0.954,1.051],["Jan-26",42.5,40.5,38.5,-2.0,2.0,0.953,1.052],["Feb-26",54.0,51.5,49.0,-2.5,2.5,0.954,1.051],["Mar-26",66.5,63.5,60.5,-3.0,3.0,0.955,1.050],["Apr-26",80.0,76.5,73.0,-3.5,3.5,0.956,1.048],["May-26",95.0,91.0,86.5,-4.0,4.5,0.958,1.052],["Jun-26",110.0,105.5,101.0,-4.5,4.5,0.959,1.045]]
for mi,row in enumerate(months_data):
    for ci,val in enumerate(row,1):
        cell=dat(ws7,r+mi,ci,val,alt=(mi%2==1))
        if ci in [2,3,4,5,6]: cell.number_format='#,##0.0'
        if ci in [7,8]: cell.number_format='0.000'
r+=14; hdr(ws7,r,1,"BUDGET BREAKDOWN BY PHASE",span=11); r+=1
for i,h in enumerate(["Phase","Budget","Committed","Spent","Remaining","% Spent","EAC","Variance","Risk Reserve"],1): hdr(ws7,r,i,h,sfill)
r+=1
for bi,row in enumerate([["Phase 1","42.5M","42.5M","38.2M","4.3M","90%","40.5M","+2.0M","2.1M"],["Phase 2","170M","110M","95M","75M","56%","165M","+5.0M","8.5M"],["Phase 3","340M","0","0","340M","0%","340M","0","17.0M"],["Phase 4","127.5M","0","0","127.5M","0%","125M","+2.5M","6.4M"],["Phase 5","170M","0","0","170M","0%","170M","0","8.5M"],["TOTAL","850M","152.5M","133.2M","716.8M","16%","840.5M","+9.5M","42.5M"]]):
    for ci,val in enumerate(row,1): dat(ws7,r+bi,ci,val,alt=(bi%2==1),bold=(bi==5))
sw(ws7,[22,14,14,14,14,10,14,14,14]); ws7.freeze_panes="A4"
print("Sheet 7: KPI & PMP DONE")

# ===== SHEET 8: RISK REGISTER =====
ws8=wb.create_sheet("Risk Register")
ws8.merge_cells('A1:K1'); ws8['A1']="PROJECT RISK REGISTER"
ws8['A1'].font=Font(bold=True,size=14,color="FFFFFF",name="Calibri"); ws8['A1'].fill=PatternFill(start_color="C00000",fill_type="solid"); ws8['A1'].alignment=ctr
r=3
for i,h in enumerate(["Risk ID","Description","Category","Prob (1-5)","Impact (1-5)","Score","Mitigation","Owner","Status","Phase","Updated"],1):
    hdr(ws8,r,i,h,PatternFill(start_color="C00000",fill_type="solid"))
r+=1
for ri,row in enumerate([["R-001","Radar delivery delayed","Schedule","3","4","12","Dual-source","Procurement","Active","Phase 2","2026-05-01"],["R-002","Site access restrictions","Schedule","2","5","10","Early coordination","Site Mgr","Active","Phase 3","2026-04-15"],["R-003","Currency fluctuation","Financial","3","3","9","Forward contracts","Finance","Active","Phase 2","2026-05-10"],["R-004","Key personnel unavail","Resource","2","4","8","Cross-training","HR","Monitoring","All","2026-03-20"],["R-005","Regulatory approval delay","Schedule","3","5","15","Early engagement","PM","Active","Phase 2","2026-05-01"],["R-006","Legacy CAD integration","Technical","4","4","16","Early prototyping","SW Arch","Active","Phase 2-3","2026-04-01"],["R-007","Underground utilities","Technical","2","4","8","GPR survey","Civil","Monitoring","Phase 3","2026-04-20"],["R-008","Weather delays","Schedule","3","3","9","Schedule buffer","Site Mgr","Monitoring","Phase 3","2026-03-15"],["R-009","Cybersecurity cert delay","Technical","2","4","8","Early CERT engagement","IT Security","Monitoring","Phase 4","2026-05-05"],["R-010","Stakeholder opposition","Stakeholder","1","3","3","Public consultation","PM","Closed","Phase 1","2025-09-10"]]):
    for ci,val in enumerate(row,1):
        cell=dat(ws8,r+ri,ci,val,alt=(ri%2==1))
        if ci==6:
            score=int(val) if str(val).isdigit() else 0
            if score>=12: cell.font=Font(bold=True,color="C00000",size=10)
            elif score>=8: cell.font=Font(color="ED7D31",size=10)
sw(ws8,[10,35,14,10,12,8,30,16,12,12,12]); ws8.freeze_panes="A4"
print("Sheet 8: Risk Register DONE")

# ===== SHEET 9: ISSUE LOG =====
ws9=wb.create_sheet("Issue Log")
ws9.merge_cells('A1:J1'); ws9['A1']="PROJECT ISSUE LOG"
ws9['A1'].font=Font(bold=True,size=14,color="FFFFFF",name="Calibri"); ws9['A1'].fill=PatternFill(start_color="7030A0",fill_type="solid"); ws9['A1'].alignment=ctr
r=3
for i,h in enumerate(["Issue ID","Description","Raised Date","By","Priority","Status","Assigned To","Due Date","Resolution","Phase"],1):
    hdr(ws9,r,i,h,PatternFill(start_color="7030A0",fill_type="solid"))
r+=1
for ii,row in enumerate([["ISS-001","SSR vendor criteria disputed","2026-04-10","PM","High","In Progress","Procurement","2026-05-30","Clarification meeting","Phase 2"],["ISS-002","Site B access road inadequate","2026-03-25","Site Eng","Medium","Open","Civil Works","2026-06-15","Road upgrade quote","Phase 2"],["ISS-003","Radar frequency conflict","2026-05-01","Comms Eng","High","In Progress","Tech Lead","2026-05-20","OFCA coordination","Phase 2"],["ISS-004","CAD ops unavailable for OAT","2026-04-20","QA","Medium","Open","PM","2026-06-30","Schedule meeting","Phase 4"],["ISS-005","PKG-001 budget overrun risk","2026-05-15","Cost Eng","High","Under Review","Finance","2026-06-15","Value engineering","Phase 2"]]):
    for ci,val in enumerate(row,1):
        cell=dat(ws9,r+ii,ci,val,alt=(ii%2==1))
        if ci==5 and val=="High": cell.font=Font(bold=True,color="C00000",size=10)
sw(ws9,[12,40,12,12,10,14,14,12,30,10]); ws9.freeze_panes="A4"
print("Sheet 9: Issue Log DONE")

# ===== SHEET 10: CHANGE REQUESTS =====
ws10=wb.create_sheet("Change Requests")
ws10.merge_cells('A1:J1'); ws10['A1']="CHANGE REQUEST LOG"
ws10['A1'].font=Font(bold=True,size=14,color="FFFFFF",name="Calibri"); ws10['A1'].fill=PatternFill(start_color="375623",fill_type="solid"); ws10['A1'].alignment=ctr
r=3
for i,h in enumerate(["CR ID","Description","Requestor","Submitted","Cost Impact","Schedule Impact","Priority","Status","Decision Date","Approved By"],1):
    hdr(ws10,r,i,h,PatternFill(start_color="375623",fill_type="solid"))
r+=1
for ci,row in enumerate([["CR-001","Add 5th MLAT station","CAD Ops","2026-04-01","+12M","+45 days","Medium","Approved","2026-05-10","Steering Comm"],["CR-002","Upgrade display to 4K","ATCO Union","2026-03-15","+3.5M","+14 days","Low","Approved","2026-04-20","PM"],["CR-003","Cybersecurity requirements","IT Security","2026-05-05","+8M","+30 days","High","Under Review","—","—"],["CR-004","Extend parallel running","CAD Ops","2026-04-20","+5M","+90 days","Medium","Pending","—","—"]]):
    for cii,val in enumerate(row,1): dat(ws10,r+ci,cii,val,alt=(ci%2==1))
sw(ws10,[10,40,14,14,12,14,10,14,14,20]); ws10.freeze_panes="A4"
print("Sheet 10: Change Requests DONE")

# Save
output = r"C:\Users\TonyChoi\Projects\RIM CAP\ExpressTire\HK_CAD_Radar_Project_Monitoring.xlsx"
wb.save(output)
print(f"\n✅ EXCEL WORKBOOK SAVED: {output}")

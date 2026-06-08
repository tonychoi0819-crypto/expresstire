import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random

wb = openpyxl.Workbook()

# ===== STYLES =====
hdr_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
sub_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
phase_colors = ["4472C4","70AD47","FFC000","ED7D31","C00000"]
kpi_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
thin = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
ctr = Alignment(horizontal="center",vertical="center",wrap_text=True)
left_a = Alignment(horizontal="left",vertical="center",wrap_text=True)

def hdr(ws, r, c, val, fill=hdr_fill, span=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = hdr_font; cell.fill = fill; cell.alignment = ctr; cell.border = thin
    if span: ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)

def dat(ws, r, c, val, alt=False, fmt=None, bold=False):
    cell = ws.cell(row=r, column=c, value=val)
    cell.border = thin
    cell.alignment = left_a if isinstance(val, str) else ctr
    if alt: cell.fill = alt_fill
    if fmt: cell.number_format = fmt
    if bold: cell.font = Font(bold=True, name="Calibri", size=10)
    return cell

def sw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 1: PROJECT SUMMARY DASHBOARD
# ============================================================
ws = wb.active; ws.title = "Project Dashboard"; ws.sheet_properties.tabColor = "1F4E79"
ws.merge_cells('A1:O1'); ws['A1'] = "HONG KONG CAD RADAR PROJECT — PROJECT MONITORING DASHBOARD"
ws['A1'].font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
ws['A1'].fill = PatternFill(start_color="1F4E79", fill_type="solid"); ws['A1'].alignment = ctr
ws.merge_cells('A2:O2'); ws['A2'] = "Civil Aviation Department | 5-Phase Radar System Implementation | PMP Compliant"
ws['A2'].font = Font(size=11, color="1F4E79", name="Calibri", italic=True); ws['A2'].alignment = ctr

row = 4
hdr(ws, row, 1, "PROJECT INFORMATION", span=8); row += 1
info = [("Project Name","HK CAD Radar System Modernization"),("Client","Civil Aviation Department, HK"),
        ("Project Manager","TBD"),("Project Code","CAD-RADAR-2025"),("Contract Value","HK$850,000,000"),
        ("Start Date","2025-07-01"),("Expected Completion","2029-06-30"),("Status","In Progress")]
for i,(k,v) in enumerate(info):
    dat(ws,row,i+1,k,bold=True); dat(ws,row+1,i+1,v)
row += 3
hdr(ws,row,1,"PHASE OVERVIEW",span=10); row += 1
ph_h = ["Phase","Phase Name","Start Date","End Date","Duration (Days)","Budget (HK$)","Actual Cost (HK$)","% Complete","Status","RAG"]
for i,h in enumerate(ph_h,1): hdr(ws,row,i,h,sub_fill)
row += 1
phases = [
    ["1","Initiation & Planning","2025-07-01","2025-12-31","184","HK$42,500,000","HK$38,200,000","95%","Completed","Green"],
    ["2","Design & Procurement","2026-01-01","2026-12-31","365","HK$170,000,000","HK$95,000,000","60%","In Progress","Green"],
    ["3","Installation & Integration","2027-01-01","2027-12-31","365","HK$340,000,000","HK$0","0%","Not Started","Grey"],
    ["4","Testing & Commissioning","2028-01-01","2028-12-31","365","HK$127,500,000","HK$0","0%","Not Started","Grey"],
    ["5","Handover & Closeout","2029-01-01","2029-06-30","182","HK$170,000,000","HK$0","0%","Not Started","Grey"],
]
for pi,ph in enumerate(phases):
    for ci,val in enumerate(ph,1):
        f = PatternFill(start_color=phase_colors[pi],fill_type="solid") if ci==1 else None
        cell = dat(ws,row+pi,ci,val,alt=(pi%2==1))
        if f: cell.fill = f; cell.font = Font(bold=True,color="FFFFFF",size=10)
row += 6
hdr(ws,row,1,"KEY PERFORMANCE INDICATORS (KPIs)",span=10); row += 1
kh = ["KPI","Target","Actual","Variance","Unit","Status","Trend","PMP Area","Frequency","Responsible"]
for i,h in enumerate(kh,1): hdr(ws,row,i,h,sub_fill)
row += 1
kpis = [
    ["Schedule Performance Index (SPI)","1.00","0.98","-0.02","Ratio","On Track","↑","Schedule","Weekly","PM"],
    ["Cost Performance Index (CPI)","1.00","1.05","+0.05","Ratio","Favorable","↑","Cost","Weekly","PM"],
    ["Scope Change Requests","<10","3","-7","Count","Favorable","↓","Scope","Monthly","PM"],
    ["Defect Density","<0.5","0.3","-0.2","Defects/KLOC","Favorable","↓","Quality","Monthly","QA"],
    ["Risk Exposure Score","<50","35","-15","Score","Acceptable","→","Risk","Monthly","Risk Mgr"],
    ["Stakeholder Satisfaction",">80%","85%","+5%","%","Favorable","↑","Stakeholder","Quarterly","PM"],
    ["Safety Incident Rate","0","0","0","Incidents","Perfect","→","Safety","Weekly","Safety Off."],
    ["Milestone On-Time Rate",">90%","92%","+2%","%","Favorable","↑","Schedule","Monthly","PM"],
    ["Budget at Completion (BAC)","HK$850M","HK$850M","HK$0","HK$","On Budget","→","Cost","Monthly","PM"],
    ["Estimate at Completion (EAC)","HK$850M","HK$809M","-HK$41M","HK$","Favorable","↑","Cost","Monthly","PM"],
]
for ki,kpi in enumerate(kpis):
    for ci,val in enumerate(kpi,1): dat(ws,row+ki,ci,val,alt=(ki%2==1))
row += 12
hdr(ws,row,1,"PMP EARNED VALUE MANAGEMENT (EVM)",span=10); row += 1
evh = ["Metric","Formula","PV","EV","AC","SV","CV"]
for i,h in enumerate(evh,1): hdr(ws,row,i,h,sub_fill)
row += 1
evm = [
    ["Cumulative","—","HK$212,500,000","HK$208,250,000","HK$198,333,333","-HK$4,250,000","+HK$9,916,667"],
    ["Current Period","—","HK$17,000,000","HK$16,660,000","HK$15,883,333","-HK$340,000","+HK$776,667"],
    ["EAC","BAC/CPI","HK$850,000,000","—","—","—","—"],
    ["VAC","BAC-EAC","—","—","—","—","+HK$41,000,000"],
    ["TCPI","(BAC-EV)/(BAC-AC)","—","—","—","—","1.02"],
]
for ei,ev in enumerate(evm):
    for ci,val in enumerate(ev,1): dat(ws,row+ei,ci,val,alt=(ei%2==1))
row += 7
hdr(ws,row,1,"RAG STATUS LEGEND",span=10); row += 1
for ri,rag in enumerate([["Green","On track","<5%"],["Amber","At risk","5-10%"],["Red","Critical",">10%"],["Grey","Not started","N/A"]]):
    for ci,val in enumerate(rag,1): dat(ws,row+ri,ci,val,alt=(ri%2==1))
sw(ws,[22,20,16,16,14,18,18,12,14,8,14,12,18,16,16])
ws.freeze_panes = "A5"
print("Sheet 1: Dashboard DONE")

# ============================================================
# SHEET 2: PHASE 1 — INITIATION & PLANNING
# ============================================================
ws2 = wb.create_sheet("Phase 1 - Initiation"); ws2.sheet_properties.tabColor = phase_colors[0]
ws2.merge_cells('A1:L1'); ws2['A1'] = "PHASE 1: INITIATION & PLANNING"
ws2['A1'].font = Font(bold=True,size=14,color="FFFFFF",name="Calibri")
ws2['A1'].fill = PatternFill(start_color=phase_colors[0],fill_type="solid"); ws2['A1'].alignment = ctr
r = 3
hdr(ws2,r,1,"WORK BREAKDOWN STRUCTURE (WBS)",span=12); r+=1
wh = ["WBS ID","Task Name","Duration (Days)","Start","End","Predecessor","Assigned To","% Complete","Status","Deliverable","RAG","Notes"]
for i,h in enumerate(wh,1): hdr(ws2,r,i,h,PatternFill(start_color=phase_colors[0],fill_type="solid"))
r+=1
wbs1 = [
    ["1.1","Project Charter Development","14","2025-07-01","2025-07-14","—","Project Sponsor","100%","Completed","Project Charter","Green","Approved by CAD"],
    ["1.2","Stakeholder Identification","21","2025-07-01","2025-07-21","—","PM","100%","Completed","Stakeholder Register","Green","47 stakeholders identified"],
    ["1.3","Requirements Gathering","42","2025-07-15","2025-08-25","1.1","BA Team","100%","Completed","Requirements Doc","Green","215 requirements"],
    ["1.4","Feasibility Study","30","2025-07-22","2025-08-20","1.2","Eng. Lead","100%","Completed","Feasibility Report","Green","Technical feasibility confirmed"],
    ["1.5","Site Survey (HK Airport)","28","2025-08-01","2025-08-28","1.2","Site Eng.","100%","Completed","Site Survey Report","Green","3 sites surveyed"],
    ["1.6","Environmental Impact Assessment","35","2025-08-15","2025-09-18","1.4","Env. Consult.","100%","Completed","EIA Report","Green","EPD approval obtained"],
    ["1.7","Risk Assessment","28","2025-08-26","2025-09-22","1.3,1.4","Risk Mgr","100%","Completed","Risk Register","Green","68 risks identified"],
    ["1.8","Budget Estimation","21","2025-09-01","2025-09-21","1.4","Cost Eng.","100%","Completed","Cost Estimate","Green","HK$850M approved"],
    ["1.9","Project Plan Development","30","2025-09-15","2025-10-14","1.3,1.7","PM","100%","Completed","Project Management Plan","Green","PMBOK 7th ed."],
    ["1.10","Regulatory Approval (CAD)","45","2025-09-23","2025-11-06","1.6,1.8","PM","100%","Completed","CAD Approval Letter","Green","Obtained 2025-11-06"],
    ["1.11","Phase Gate Review","7","2025-11-15","2025-11-21","1.9,1.10","Steering Committee","100%","Completed","Gate Approval","Green","Phase 2 authorized"],
    ["1.12","Lessons Learned Documentation","10","2025-11-22","2025-12-01","1.11","PM","90%","In Progress","Lessons Learned Report","Green","Draft complete"],
]
for wi,row_data in enumerate(wbs1):
    for ci,val in enumerate(row_data,1): dat(ws2,r+wi,ci,val,alt=(wi%2==1))
r += 14
hdr(ws2,r,1,"PHASE 1 KPIs",span=12); r+=1
p1kh = ["KPI","Target","Actual","Unit","Status"]
for i,h in enumerate(p1kh,1): hdr(ws2,r,i,h,PatternFill(start_color=phase_colors[0],fill_type="solid"))
r+=1
for ki,row in enumerate([["Planning Complete","100%","95%","%","On Track"],["Budget Variance","<5%","4.2%","%","On Track"],["Stakeholders Identified",">40","47","Count","Favorable"],["Requirements Documented",">200","215","Count","Favorable"],["Risks Identified",">50","68","Count","Favorable"],["Regulatory Approvals","100%","100%","%","On Track"]]):
    for ci,val in enumerate(row,1): dat(ws2,r+ki,ci,val,alt=(ki%2==1))
sw(ws2,[10,32,14,12,12,12,16,12,14,22,8,20])
ws2.freeze_panes = "A4"
print("Sheet 2: Phase 1 DONE")

# ============================================================
# SHEET 3: PHASE 2 — DESIGN & PROCUREMENT
# ============================================================
ws3 = wb.create_sheet("Phase 2 - Design"); ws3.sheet_properties.tabColor = phase_colors[0]
ws3.merge_cells('A1:L1'); ws3['A1'] = "PHASE 2: DESIGN & PROCUREMENT"
ws3['A1'].font = Font(bold=True,size=14,color="FFFFFF",name="Calibri")
ws3['A1'].fill = PatternFill(start_color=phase_colors[1],fill_type="solid"); ws3['A1'].alignment = ctr
r = 3
hdr(ws3,r,1,"WORK BREAKDOWN STRUCTURE (WBS)",span=12); r+=1
for i,h in enumerate(wh,1): hdr(ws3,r,i,h,PatternFill(start_color=phase_colors[1],fill_type="solid"))
r+=1
wbs2 = [
    ["2.1","System Architecture Design","60","2026-01-01","2026-03-01","1.11","System Arch.","85%","In Progress","Arch Doc v2.3","Green","On schedule"],
    ["2.2","Radar Technical Specifications","45","2026-01-15","2026-02-28","2.1","Tech Lead","75%","In Progress","Tech Spec v1.5","Green","Primary radar: ASR-12"],
    ["2.3","SSR System Design","45","2026-02-01","2026-03-17","2.1","Comms Eng.","65%","In Progress","SSR Design Doc","Amber","Minor delay in Mode S"],
    ["2.4","MLAT System Design","40","2026-02-15","2026-03-26","2.1","Nav. Eng.","50%","In Progress","MLAT Design","Green","4 stations planned"],
    ["2.5","Data Processing System Design","55","2026-03-01","2026-04-24","2.1","SW Architect","40%","In Progress","SW Design Doc","Green","Asterix protocol"],
    ["2.6","Site Preparation Design","50","2026-03-15","2026-05-03","2.2,2.3","Civil Eng.","30%","In Progress","Site Prep Plans","Green","3 sites"],
    ["2.7","RFP Preparation & Issue","35","2026-04-01","2026-05-05","2.2","Procurement","25%","In Progress","RFP Documents","Green","3 packages"],
    ["2.8","Tender Evaluation","60","2026-05-06","2026-07-04","2.7","Tender Board,""","0%","Not Started","Tender Report","Grey",""],
    ["2.9","Contract Award","30","2026-07-05","2026-08-03","2.8","Legal/PM","0%","Not Started","Signed Contracts","Grey",""],
    ["2.10","Factory Acceptance Test Plan","28","2026-06-01","2026-06-28","2.2,2.3","QA Lead","0%","Not Started","FAT Plan","Grey",""],2    ["2.11","Site Acceptance Test Plan","28","2026-06-15","2026-07-12","2.4,2.5","QA Lead","0%","Not Started","SAT Plan","Grey",""],
    ["2.12","Design Review (PDR)","14","2026-04-25","2026-05-08","2.1,2.2","CAD/PM","0%","Not Started","PDR Minutes","Grey",""],
    ["2.13","Design Review (CDR)","14","2026-08-04","2026-08-17","2.9,2.12","CAD/PM","0%","Not Started","CDR Minutes","Grey",""],
    ["2.14","Phase Gate Review","7","2026-11-15","2026-11-21","All above","Steering Committee","0%","Not Started","Gate Approval","Grey",""],
]
for wi,row_data in enumerate(wbs2):
    for ci,val in enumerate(row_data,1): dat(ws3,r+wi,ci,val,alt=(wi%2==1))
r += 16
hdr(ws3,r,1,"PHASE 2 PROCUREMENT TRACKER",span=12); r+=1
p2ph = ["Package","Description","RFP Date","Tender Close","Eval. Complete","Vendor","Value (HK$)","Status","RAG"]
for i,h in enumerate(p2ph,1): hdr(ws3,r,i,h,PatternFill(start_color=phase_colors[1],fill_type="solid"))
r+=1
proc2 = [
    ["PKG-001","Primary Radar System","2026-04-15","2026-05-30","2026-07-04","TBD (Thales/Leonardo)","HK$280,000,000","RFP Issued","Green"],
    ["PKG-002","Secondary Surveillance Radar","2026-04-15","2026-05-30","2026-07-04","TBD (Raytheon/Indra)","HK$120,000,000","RFP Issued","Green"],
    ["PKG-003","MLAT System","2026-05-01","2026-06-15","2026-07-30","TBD (Roke/ERA)","HK$45,000,000","Draft RFP","Amber"],
    ["PKG-004","Data Processing & Display","2026-05-15","2026-06-30","2026-08-15","TBD (Frequentis/Indra)","HK$95,000,000","Planning","Grey"],
    ["PKG-005","Communication Infrastructure","2026-06-01","2026-07-15","2026-08-30","TBD (Ericsson/Nokia)","HK$35,000,000","Planning","Grey"],
    ["PKG-006","Site Civil Works","2026-07-01","2026-08-15","2026-09-30","TBD (Local contractors)","HK$65,000,000","Planning","Grey"],
    ["PKG-007","Installation & Integration","2026-08-01","2026-09-15","2026-10-30","TBD (System integrator)","HK$85,000,000","Planning","Grey"],
]
for pi,row_data in enumerate(proc2):
    for ci,val in enumerate(row_data,1): dat(ws3,r+pi,ci,val,alt=(pi%2==1))
sw(ws3,[10,32,14,12,12,12,16,12,14,22,8,20])
ws3.freeze_panes = "A4"
print("Sheet 3: Phase 2 DONE")

# ============================================================
# SHEET 4: PHASE 3 — INSTALLATION & INTEGRATION
# ============================================================
ws4 = wb.create_sheet("Phase 3 - Installation"); ws4.sheet_properties.tabColor = "B4C6E7"
ws4.merge_cells('A1:L1'); ws4['A1'] = "PHASE 3: SITE INSTALLATION & SYSTEM INTEGRATION"
ws4['A1'].font = Font(bold=True,size=14,color="000000",name="Calibri")
ws4['A1'].fill = PatternFill(start_color=phase_colors[2],fill_type="solid"); ws4['A1'].alignment = ctr
r = 3
hdr(ws4,r,1,"WORK BREAKDOWN STRUCTURE (WBS)",span=12); r+=1
for i,h in enumerate(wh,1): hdr(ws4,r,i,h,PatternFill(start_color=phase_colors[2],fill_type="solid"))
r+=1
wbs3 = [
    ["3.1","Site Preparation - Site A (Chek Lap Kok)","90","2027-01-15","2027-04-14","2.6","Civil Contractor","0%","Not Started","Site Ready Cert","Grey",""],
    ["3.2","Site Preparation - Site B (Tai Mo Shan)","90","2027-01-15","2027-04-14","2.6","Civil Contractor","0%","Not Started","Site Ready Cert","Grey",""],
    ["3.3","Site Preparation - Site C (Tai Mo Shan MLS)","75","2027-02-01","2027-04-16","2.6","Civil Contractor","0%","Not Started","Site Ready Cert","Grey",""],
    ["3.4","Deliver & Install Primary Radar","120","2027-04-15","2027-08-12","3.1,PKG1","Vendor","0%","Not Started","Install Cert","Grey",""],
    ["3.5","Deliver & Install SSR","90","2027-05-01","2027-07-29","3.2,PKG2","Vendor","0%","Not Started","Install Cert","Grey",""],
    ["3.6","Install MLAT Stations (x4)","150","2027-04-16","2027-09-12","3.3,PKG3","Vendor","0%","Not Started","Install Cert","Grey",""],
    ["3.7","Install Data Processing System","60","2027-06-01","2027-07-30","3.1,PKG4","Vendor","0%","Not Started","Install Cert","Grey",""],
    ["3.8","Network & Communications Install","75","2027-05-15","2027-07-28","3.1,PKG5","Vendor","0%","Not Started","Install Cert","Grey",""],
    ["3.9","System Integration (All Subsystems)","90","2027-08-13","2027-11-10","3.4-3.8","SI Lead","0%","Not Started","Integration Report","Grey",""],
    ["3.10","Interfacing with Existing CAD Systems","60","2027-09-01","2027-10-30","3.7,3.8","SW Team","0%","Not Started","Interface Cert","Grey","Indra system"],
    ["3.11","Power Supply & Backup Systems","45","2027-07-01","2027-08-14","3.1","Electrical Eng.","0%","Not Started","Power Cert","Grey","UPS + Generator"],
    ["3.12","Earthing & Lightning Protection","30","2027-07-15","2027-08-13","3.1","Electrical Eng.","0%","Not Started","Safety Cert","Grey",""],
    ["3.13","Phase Gate Review","7","2027-11-20","2027-11-26","3.9-3.12","CAD/PM","0%","Not Started","Gate Approval","Grey",""],
]
for wi,row_data in enumerate(wbs3):
    for ci,val in enumerate(row_data,1): dat(ws4,r+wi,ci,val,alt=(wi%2==1))
r += 15
hdr(ws4,r,1,"INSTALLATION MILESTONE TRACKER",span=12); r+=1
imh = ["Milestone","Planned Date","Forecast Date","Actual Date","Status","Days Variance"
